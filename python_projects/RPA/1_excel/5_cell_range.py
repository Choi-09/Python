from openpyxl import Workbook
from random import *

wb= Workbook()
ws = wb.active

# 한 줄씩 데이터 넣기
ws.append(['번호', '영어', '수학']) # 번호: A컬럼, 영어: B컬럼, 수학: C컬럼
for i in range(1,11):
    ws.append([i, randint(0,100), randint(0,100) ])


col_B = ws['B'] # 영어 컬럼만 가져온다 
print(col_B)

# B 컬럼 셀의 값 출력
for cell in col_B:
    print(cell.value)

print("===============================")

# 영어, 수학 컬럼 함께 출력
col_range = ws['B:C']
for cols in col_range:
    for cell in cols:
        print(cell.value)
print("===============================")

# row만 가져오기
row_title = ws[1] # 첫 번째 row만 가져 오기
for cell in row_title:
    print(cell.value)

print("===============================")

row_range = ws[2:6]  # 주의! 1번째 줄인 title 제외 하고 2~6번째 줄 까지 포함! 
for rows in row_range:
    for cell in rows:
        print(cell.value, end = " ")
    print()

print("===============================")

from openpyxl.utils.cell import coordinate_from_string

# row 범위를 모를 때
row_range = ws[2 : ws.max_row] # -1 안먹힌다 
for rows in row_range:
    for cell in rows:
        print(cell.value, end = " ")
        print(cell.coordinate, end = " ") # 각 셀의 좌표 정보를 보여준다. 
        xy = coordinate_from_string(cell.coordinate) # 튜플 형태로 셀 좌표정보를 보여준다.
        print(xy, end = " ")
        print(xy[0], end = "") # == A
        print(xy[1], end = " ")  # == 1 
    print()

print("===============================")

# 전체 rows 가져오기
print(tuple(ws.rows)) # 한 row를 한 tuple로 묶는다
for row in tuple(ws.rows):
    print(row[1].value) # row 인덱스 1의 값을 전체 row를 돌면서 가져온다 => column 인덱스 1의 값과 동일 

for row in ws.iter_rows(): # 전체 row
    print(row[1].value)

for row in ws.iter_rows(min_row = 1, max_row=5): # 1번째 row ~ 5번째 row 까지
    print(row[2].value)

for row in ws.iter_rows(min_row = 2, max_row = 11, min_col = 2, max_col = 3): # 2번째 row ~ 11번째 row 까지 & 2번째 col ~ 3번째 col 까지
    print(row[0].value, row[1].value) # 수학 row, 영어 row  ∵ 2번째 컬럼부터 3번째 컬럼만 지정했으므로 row[0]이 수학, row[1]이 영어 
    print(row) # 헷갈리면 row 를 출력해보자

print("===============================")

# 전체 colums 가져오기
print(tuple(ws.columns)) # 한 column을 한 tuple로 묶는다.
for column in tuple(ws.columns):
    print(column[0].value) # column 인덱스 0의 값을 전체 column을 돌면서 가져온다 => row 인덱스 0의 값과 동일 

for col in ws.iter_cols(): # 전체 col
    print(col[0].value)

for col in ws.iter_cols(min_row = 1, max_row = 5, min_col = 1, max_col=3): # 1~5번째 row, 1~3번째 col, 지정하지 않으면 최소, 최대 값을 자동으로 지정한다. 필요한 부분만 적으면 됨
    print(col)
    print(col[0].value, col[0])

wb.save("rpa_test5.xlsx")