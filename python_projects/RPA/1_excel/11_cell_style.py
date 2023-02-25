from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook

wb = load_workbook("rpa_test5.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws['A1'] # 번호
b1 = ws['B1'] # 영어
c1 = ws["C1"] # 수학

# 셀 너비 지정
ws.column_dimensions['A'].width = 5

# 셀 높이 지정
ws.row_dimensions[1].height = 50

# 스타일 적용
# 1. 폰트 적용: from openpyxl.styles import Font
a1.font = Font(color = "FF0000", italic = True, bold = True) # 색상 RGB 빨강, 이탤릭체 적용, 볼드 적용
b1.font = Font(color = "CC33FF", name = "Arial", strike = True) # 색상 RGB CC33FF, 글꼴 Arial, 취소선(strike) 적용
c1.font = Font(color = "0000FF", size = 20, underline="double") # 색상 RGB 0000FF, 글자크기 20, 밑줄 double 적용(single / double 옵션)

# 2. 테두리 적용: from openpyxl.styles import Border, Side
thin_border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom= Side(style = "thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 3. 배경색 적용
# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:

# 4. cell 정렬: from openpyxl.styles import Alignment
        cell.alignment = Alignment(horizontal="center", vertical= "center") # center, left, right, top, bottom
        if cell == 1: # A 번호 열은 제외
            continue
        
        if isinstance(cell.value, int) and cell.value > 90: # isinstance: cell의 값이 정수형이고 90 보다 높으면
            # from openpyxl.styles import PatternFill
            cell.fill = PatternFill(fgColor = "00FF00", fill_type= "solid") # 배경색 설정
            cell.font = Font(color="FF0000") # 폰트 색상 변경

# 5. 틀 고정
ws.freeze_panes = "B2" # B2기준으로 틀 고정

wb.save("rpa_test5_style.xlsx")