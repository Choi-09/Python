from openpyxl import load_workbook

wb = load_workbook("rpa_test.xlsx") # rpa_test.xlsx 파일에서 wb를 불러옴
ws = wb.active  # 활성화 된 sheet

# cell 데이터 불러오기
# cell 갯수를 알 때
for x in range(1,11):
    for y in range(1,5):
        print(ws.cell(row = x, column = y).value, end = " ") # 줄바꿈 방지를 위해 end 공백 설정
    
    print()

print("========================================")

# cell 갯수를 모를 때
for x in range(1,ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(row = x, column = y).value, end = " ")
    
    print()
