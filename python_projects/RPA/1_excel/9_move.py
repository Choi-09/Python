from openpyxl import load_workbook

wb = load_workbook("rpa_test5.xlsx")
ws = wb.active

# 현재 엑셀파일: 번호, 영어, 수학
# 이동 후 엑셀파일: 번호, 국어, 영어, 수학
# ws.move_range("B1:C11", rows=0, cols=1)  # (이동하려는 범위 지정, 얼마나 옮길지 지정) 
ws.move_range("C1:C11", rows = 0, cols = -1) # 마이너스로 범위 지정하면 원래있던 값은 사라진다. 
# ws['B1'].value = '국어'
wb.save("rpa_test5_move_cols.xlsx")