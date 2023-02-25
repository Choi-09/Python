from openpyxl import load_workbook

wb = load_workbook("rpa_test5.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row = 2): # 제일 첫 row(컬럼명)은 제외
    # 번호, 영어, 수학 순
    if int(row[1].value) > 90:
       print(row[0].value, "번 학생은 영어 천재")
    if int(row[2].value) > 90:
        print(row[0].value, "번 학생은 수학 천재")

for row in ws.iter_rows(max_row = 1): # 첫 번째 row만 지정
    for cell in row:
        if (cell.value == '영어'):
            cell.value = '컴퓨터' # '영어' 컬럼명을 '컴퓨터'로 변경

wb.save("rpa_test5_modified.xlsx")