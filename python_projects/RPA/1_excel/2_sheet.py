from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 처음 활성화 된 sheet 뒤에 기본 이름으로 새로운 sheet 추가 생성

ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff9966" # RGB 형태로 값을 넣어 주면 탭 색상 변경

# Sheet, MySheet, YourSheet 
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("OurSheet", 2) # 2번째 인덱스에 sheet 생성

new_ws = wb['OurSheet'] # Dict 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"  # A1셀의 값을 Test
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("rpa_test2.xlsx")

