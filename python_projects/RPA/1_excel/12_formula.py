import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
ws['A2'] = "=SUM(1, 2, 3)" # 1+2+3 = 6 
ws['A3'] = "=Average(1,2,3)" # (1+2+3)/3 = 2

ws['A4'] = 10
ws['A5'] = 20
ws['A6'] = 30
ws["A7"] = "=SUM(A4:A6)"


wb.save("rpa_formula.xlsx")