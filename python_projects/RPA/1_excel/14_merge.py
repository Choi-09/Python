from openpyxl import Workbook
from openpyxl import load_workbook

# 병합하기
wb = Workbook()
ws = wb.active

ws.merge_cells("B2:D2")  # B2부터 D2까지 병합
ws['B2'].value = "Merged Cell"
wb.save("rpa_merge.xlsx")


# 셀 나누기
wb2 = load_workbook("rpa_merge.xlsx")
ws2 = wb2.active

ws2.unmerge_cells("B2:D2") # B2:D2 까지 병합되어 있던 셀을 다시 나누기
wb2.save("rpa_unmerge.xlsx")



