from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "TestSheet"

# 각 셀에 값 입력하기 
ws["A1"] = 1  # A1셀에 1 
ws['A2'] = 2  # A2셀에 2
ws['A3'] = 3  # A3셀에 3
ws['A4'] = 4
ws['A5'] = 5

print(ws['A1']) # A1 셀의 정보 출력
print(ws['A1'].value) # A1 셀의 값 출력

print(ws['A10'].value) # 값이 없는 셀의 출력: None

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ... 
print(ws.cell(column = 1, row=1).value) # == ws['A1'].value


c = ws.cell(column = 3, row = 1, value = 10) # ws['C1'].value = 10 
print(c.value) # == ws['C1'].value

'''
반복문을 통해서 셀에 데이터 입력
'''
from random import *
index = 1
for x in range(1,11): # 10개 row
    for y in range(1,5):  # 4개 column
        # ws.cell(column = y, row = x, value = randint(0,100)) # 0~100 사이의 랜덤한 정수 입력 
        ws.cell(column = y, row = x, value = index)
        index +=1

wb.save("rpa_test3.xlsx")