from openpyxl import load_workbook

# wb = load_workbook("rpa_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져올 때
# for row in ws.values: # 각 셀에 대해서 값을 바로 가져온다
#     for cell in row:
#         print(cell) 

# 수식이 아닌 수식이 적용된 값을 가져올 때
wb = load_workbook("rpa_formula.xlsx", data_only=True)
ws = wb.active
for row in ws.values: # 각 셀에 대해서 값을 바로 가져온다
    for cell in row:
        print(cell) # None 이라고 뜨는 건 수식이 있는 셀 (evaluate 되지 않은 상태). 엑셀을 열 때 엑셀 안에서 수식을 계산해서 계산된 값을 보여주는것이지, 파이썬에서는 계산 해 주지 않음. 왜? 문자열로 넣었음. 그래서 엑셀 닫을 때 '저장하시겠습니까?' 창이 뜬다. yes 클릭 후 다시 run 하면 터미널에 계산된 값이 나옴. 엑셀파일은 그대로 수식있다.